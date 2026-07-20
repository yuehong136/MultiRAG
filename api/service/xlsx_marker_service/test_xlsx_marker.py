"""
xlsx_marker_service 测试集

- 合成模板（openpyxl 构造）做确定性单测：parse / recognize / 固定填充 / 动态插行 / 防御分支
- 真实样例（教职工信息登记表）做集成测试，文件不存在时自动跳过

运行: pytest api/service/xlsx_marker_service/test_xlsx_marker.py -v
"""

import os
import re
import zipfile

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from api.service.xlsx_marker_service import auto_recognize_placeholders, fill_workbook, parse_xlsx

YELLOW = "00FFFDE7"
BLUE = "00E7EEF7"

REAL_SAMPLE = os.environ.get("XLSX_MARKER_REAL_SAMPLE", "/Users/naimehao/Downloads/教职工信息登记表_AI智能填表测试.xlsx")
# WPS 另存的同一模板：sharedStrings 化、命名空间声明更多、ARGB 带 FF alpha 前缀、zip 含目录项
WPS_SAMPLE = os.environ.get("XLSX_MARKER_WPS_SAMPLE", "/Users/naimehao/Downloads/教职工信息登记表_AI智能填表测试-刻意修改后保存.xlsx")


def build_template(path: str):
    """
    合成模板（坐标均为 0-based 注释）：
    - 表单区: B2=姓名->C2待填, D2=性别->E2待填(下拉), B3=住址->C3:E3合并待填
    - 清单区: 表头 row5(B5称谓/C5姓名/D5电话), 数据区 row6-7（预留2行）
    - 下方: B9=备注标签, B10:D11 合并大格待填(summary)
    - 第二个 sheet "说明" 不参与填充
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "登记表"
    yellow = PatternFill(fill_type="solid", start_color=YELLOW, end_color=YELLOW)
    blue = PatternFill(fill_type="solid", start_color=BLUE, end_color=BLUE)

    ws["B2"] = "姓名"
    ws["B2"].fill = blue
    ws["C2"].fill = yellow
    ws["D2"] = "性别"
    ws["D2"].fill = blue
    ws["E2"].fill = yellow
    dv = DataValidation(type="list", formula1='"男,女"')
    ws.add_data_validation(dv)
    dv.add("E2")

    ws["B3"] = "住址"
    ws["B3"].fill = blue
    ws.merge_cells("C3:E3")
    ws["C3"].fill = yellow

    ws["B5"], ws["C5"], ws["D5"] = "称谓", "姓名", "电话"
    for cell in ("B5", "C5", "D5"):
        ws[cell].fill = blue
    for r in (6, 7):
        for col in "BCD":
            ws[f"{col}{r}"].fill = yellow

    ws["B9"] = "备注"
    ws["B9"].fill = blue
    ws.merge_cells("B10:D11")
    ws["B10"].fill = yellow

    ws2 = wb.create_sheet("说明")
    ws2["A1"] = "不要动这个sheet"
    wb.save(path)


PLACEHOLDERS = [
    {"path": "sheet[0]/row[1]/cell[2]", "label": "姓名", "field_key": "$1", "type": "cell"},
    {"path": "sheet[0]/row[1]/cell[4]", "label": "性别", "field_key": "$2", "type": "cell"},
    {"path": "sheet[0]/row[2]/cell[2]", "label": "住址", "field_key": "$3", "type": "cell"},
    {"path": "sheet[0]/row[9]/cell[1]", "label": "备注", "field_key": "$4", "type": "summary"},
    {
        "path": "sheet[0]/row[4]/cell[1]",
        "label": "家庭成员",
        "field_key": "$5",
        "type": "table",
        "table_config": {
            "dynamic": False,
            "header_row": 4,
            "data_start_row": 5,
            "data_end_row": 6,
            "columns": [
                {"cell_index": 1, "name": "称谓"},
                {"cell_index": 2, "name": "姓名"},
                {"cell_index": 3, "name": "电话"},
            ],
        },
    },
]


@pytest.fixture
def template(tmp_path):
    path = str(tmp_path / "template.xlsx")
    build_template(path)
    return path


# ---------- parse ----------


def test_parse_structure(template):
    wb = parse_xlsx(template, "template.xlsx")
    assert len(wb.sheets) == 2
    sheet = wb.sheets[0]
    assert sheet.name == "登记表"

    name_label = sheet.rows[1].cells[1]
    assert name_label.value == "姓名"
    assert name_label.style.fill_color == BLUE

    gender_fill = sheet.rows[1].cells[4]
    assert gender_fill.validation_options == ["男", "女"]
    assert gender_fill.style.fill_color == YELLOW

    merged_origin = sheet.rows[2].cells[2]  # C3:E3
    assert merged_origin.is_merged_origin and merged_origin.col_span == 3
    merged_member = sheet.rows[2].cells[3]
    assert not merged_member.is_merged_origin
    assert merged_member.merge_origin_path == "sheet[0]/row[2]/cell[2]"


# ---------- recognize ----------


def test_recognize_right_fill(template):
    wb = parse_xlsx(template, "template.xlsx")
    placeholders = auto_recognize_placeholders(wb)
    labels = {p.label for p in placeholders}
    assert "姓名" in labels
    assert "性别" in labels
    assert "住址" in labels
    gender = next(p for p in placeholders if p.label == "性别")
    assert gender.validation_options == ["男", "女"]
    # 说明 sheet 无候选
    assert all(p.path.startswith("sheet[0]") for p in placeholders)


def test_recognize_table_region(template):
    """清单区（连续3表头格+下方空数据行）整体识别为 dynamic_table，区域内不再产 cell 候选"""
    wb = parse_xlsx(template, "template.xlsx")
    placeholders = auto_recognize_placeholders(wb)

    tables = [p for p in placeholders if p.type == "dynamic_table"]
    assert len(tables) == 1
    table = tables[0]
    cfg = table.table_config
    assert cfg.dynamic is True
    # 表头 row4；数据区 row5-6（黄色预留）+ row7（空白分隔行也计入预留区）
    assert (cfg.header_row, cfg.data_start_row, cfg.data_end_row) == (4, 5, 7)
    assert [(c.cell_index, c.name) for c in cfg.columns] == [(1, "称谓"), (2, "姓名"), (3, "电话")]
    # 上方无标题行，回退命名；path 挂在数据区首格
    assert table.label == "表格1"
    assert table.path == "sheet[0]/row[5]/cell[1]"

    # 原先每列首行会被"下填充"命中为 cell 候选，现应全部剔除
    cell_paths = {p.path for p in placeholders if p.type == "cell"}
    assert not cell_paths & {f"sheet[0]/row[5]/cell[{c}]" for c in (1, 2, 3)}
    # 表单区候选不受影响
    assert "sheet[0]/row[1]/cell[2]" in cell_paths


def test_recognize_table_title_naming(tmp_path):
    """表头紧邻上方的标题行作为主题名称，且去掉"六、"类编号前缀"""
    path = str(tmp_path / "titled.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:D1")
    ws["A1"] = "六、主要家庭成员"
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "称谓", "姓名", "年龄", "电话"
    # row3-5 为空数据区，A6 写边界行（同时把 max_row 撑到 6）
    ws["A6"] = "备注"
    wb.save(path)

    parsed = parse_xlsx(path, "titled.xlsx")
    placeholders = auto_recognize_placeholders(parsed)
    tables = [p for p in placeholders if p.type == "dynamic_table"]
    assert len(tables) == 1
    assert tables[0].label == "主要家庭成员"
    assert (tables[0].table_config.header_row, tables[0].table_config.data_end_row) == (1, 4)


# ---------- fill: 固定区域 ----------


def test_fill_fixed(template, tmp_path):
    out = str(tmp_path / "filled.xlsx")
    fields = [
        {"id": "$1", "value": "张明华"},
        {"id": "$2", "value": "男"},
        {"id": "$3", "value": "上海市延安西路1882号"},
        {"id": "$4", "value": "第一行\n第二行"},
        {
            "id": "$5",
            "rows": [
                {"称谓": "父亲", "姓名": "张大山", "电话": "13800000000"},
                {"称谓": "母亲", "姓名": "李小花", "电话": "13900000000"},
            ],
        },
    ]
    fill_workbook(template, PLACEHOLDERS, fields, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["登记表"]
    assert ws["C2"].value == "张明华"
    assert ws["E2"].value == "男"
    assert ws["C3"].value == "上海市延安西路1882号"
    assert ws["B10"].value == "第一行\n第二行"
    assert ws["B6"].value == "父亲" and ws["D7"].value == "13900000000"
    # 下拉/合并保留
    assert len(ws.data_validations.dataValidation) == 1
    assert len(ws.merged_cells.ranges) == 2
    # 其余部件字节级一致（只允许目标 sheet XML 变化）
    with zipfile.ZipFile(template) as z1, zipfile.ZipFile(out) as z2:
        diff = [n for n in z1.namelist() if z1.read(n) != z2.read(n)]
        assert diff == ["xl/worksheets/sheet1.xml"]


def test_fill_fixed_truncates_overflow(template, tmp_path):
    out = str(tmp_path / "filled.xlsx")
    fields = [
        {
            "id": "$5",
            "rows": [
                {"称谓": "父亲", "姓名": "A", "电话": "1"},
                {"称谓": "母亲", "姓名": "B", "电话": "2"},
                {"称谓": "配偶", "姓名": "C", "电话": "3"},  # 超出预留2行，应截断
            ],
        }
    ]
    fill_workbook(template, PLACEHOLDERS, fields, out)
    ws = openpyxl.load_workbook(out)["登记表"]
    assert ws["B7"].value == "母亲"
    assert ws["B8"].value is None  # 未插行未越界


# ---------- fill: 动态插行 ----------


def test_fill_dynamic_insert(template, tmp_path):
    out = str(tmp_path / "filled.xlsx")
    dynamic_placeholders = [dict(p) for p in PLACEHOLDERS]
    dynamic_placeholders[4] = {
        **PLACEHOLDERS[4],
        "type": "dynamic_table",
        "table_config": {**PLACEHOLDERS[4]["table_config"], "dynamic": True},
    }
    fields = [
        {"id": "$4", "value": "备注内容"},
        {
            "id": "$5",
            "rows": [
                {"称谓": "父亲", "姓名": "A", "电话": "1"},
                {"称谓": "母亲", "姓名": "B", "电话": "2"},
                {"称谓": "配偶", "姓名": "C", "电话": "3"},
                {"称谓": "长子", "姓名": "D", "电话": "4"},
            ],
        },
    ]
    fill_workbook(template, dynamic_placeholders, fields, out)

    ws = openpyxl.load_workbook(out)["登记表"]
    # 4 行数据: 原 6-7 + 新插 8-9
    assert ws["B6"].value == "父亲" and ws["B9"].value == "长子"
    # 新行继承模板行样式
    assert ws["C8"].fill.start_color.rgb == YELLOW
    # 下方内容平移 2 行: B9备注标签->B11, 合并区 B10:D11 -> B12:D13
    assert ws["B11"].value == "备注"
    merges = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "B12:D13" in merges
    # summary 写入的值随平移落在 B12
    assert ws["B12"].value == "备注内容"
    # 下拉校验位置不变（在插入点上方）
    dvs = list(ws.data_validations.dataValidation)
    assert str(dvs[0].sqref) == "E2"


def test_fill_dynamic_clones_row_merges(tmp_path):
    """模板数据行内有横向合并（如 C6:D6）时，新行克隆该合并"""
    path = str(tmp_path / "t.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B5"], ws["C5"] = "项目", "说明"
    ws.merge_cells("C6:D6")
    wb.save(path)

    out = str(tmp_path / "filled.xlsx")
    placeholders = [
        {
            "path": "sheet[0]/row[4]/cell[1]",
            "label": "清单",
            "field_key": "$1",
            "type": "dynamic_table",
            "table_config": {
                "dynamic": True,
                "header_row": 4,
                "data_start_row": 5,
                "data_end_row": 5,
                "columns": [{"cell_index": 1, "name": "项目"}, {"cell_index": 2, "name": "说明"}],
            },
        }
    ]
    fields = [{"id": "$1", "rows": [{"项目": "甲", "说明": "x"}, {"项目": "乙", "说明": "y"}]}]
    fill_workbook(path, placeholders, fields, out)

    ws2 = openpyxl.load_workbook(out).active
    merges = sorted(str(r) for r in ws2.merged_cells.ranges)
    assert merges == ["C6:D6", "C7:D7"]
    assert ws2["B7"].value == "乙"


def test_fill_rejects_formula_below_insert(tmp_path):
    path = str(tmp_path / "t.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B5"] = "项目"
    ws["B6"] = None
    ws["B8"] = "=SUM(B6:B7)"
    wb.save(path)

    out = str(tmp_path / "filled.xlsx")
    placeholders = [
        {
            "path": "sheet[0]/row[4]/cell[1]",
            "label": "清单",
            "field_key": "$1",
            "type": "dynamic_table",
            "table_config": {
                "dynamic": True,
                "header_row": 4,
                "data_start_row": 5,
                "data_end_row": 5,
                "columns": [{"cell_index": 1, "name": "项目"}],
            },
        }
    ]
    fields = [{"id": "$1", "rows": [{"项目": "甲"}, {"项目": "乙"}]}]
    with pytest.raises(ValueError, match="公式"):
        fill_workbook(path, placeholders, fields, out)


def test_fill_preserves_foreign_parts(template, tmp_path):
    """zip 里塞一个引擎不认识的部件（模拟形状/自定义 XML），填充后必须原样保留"""
    hacked = str(tmp_path / "hacked.xlsx")
    with zipfile.ZipFile(template) as zin, zipfile.ZipFile(hacked, "w") as zout:
        for item in zin.namelist():
            zout.writestr(item, zin.read(item))
        zout.writestr("xl/custom/shape_like_part.xml", "<custom>盖章处</custom>")

    out = str(tmp_path / "filled.xlsx")
    fill_workbook(hacked, PLACEHOLDERS[:1], [{"id": "$1", "value": "张三"}], out)
    with zipfile.ZipFile(out) as z:
        assert z.read("xl/custom/shape_like_part.xml") == "<custom>盖章处</custom>".encode()


# ---------- 真实样例集成测试 ----------


@pytest.mark.skipif(not os.path.exists(REAL_SAMPLE), reason="真实样例文件不存在")
def test_real_sample_parse_and_fill(tmp_path):
    wb = parse_xlsx(REAL_SAMPLE, os.path.basename(REAL_SAMPLE))
    assert [s.name for s in wb.sheets] == ["教职工信息登记表", "填表说明", "AI填表评估表"]
    main = wb.sheets[0]
    # C6 待填格: 淡黄填充; E6 性别下拉
    assert main.rows[5].cells[2].style.fill_color == YELLOW
    assert main.rows[5].cells[4].validation_options == ["男", "女"]

    # recognize 在真实样例上应识别出主要表单字段
    recognized = auto_recognize_placeholders(wb)
    labels = {p.label for p in recognized}
    assert {"姓名", "性别", "身份证号"} <= labels

    # 两个清单区整体识别为 dynamic_table（主题名称取自"六、/七、"标题行并去前缀）
    rec_tables = {p.label: p for p in recognized if p.type == "dynamic_table" and p.path.startswith("sheet[0]")}
    assert {"主要家庭成员", "学习与工作经历"} <= set(rec_tables)
    family_cfg = rec_tables["主要家庭成员"].table_config
    assert (family_cfg.header_row, family_cfg.data_start_row, family_cfg.data_end_row) == (27, 28, 31)
    assert [c.name for c in family_cfg.columns] == ["称谓", "姓名", "年龄", "工作单位", "职务", "联系电话"]
    exp_cfg = rec_tables["学习与工作经历"].table_config
    assert (exp_cfg.header_row, exp_cfg.data_start_row, exp_cfg.data_end_row) == (33, 34, 38)
    # 表格区域内不再有 cell 候选（此前每列首行会被"下填充"命中）
    cell_rows = {int(re.search(r"row\[(\d+)\]", p.path).group(1)) for p in recognized if p.type == "cell" and p.path.startswith("sheet[0]")}
    assert not cell_rows & set(range(27, 39))

    # 动态填充: 家庭成员表(表头27, 数据28-31, 0-based)扩到6行, 同时填基本字段
    out = str(tmp_path / "filled.xlsx")
    placeholders = [
        {"path": "sheet[0]/row[5]/cell[2]", "label": "姓名", "field_key": "$1", "type": "cell"},
        {"path": "sheet[0]/row[5]/cell[4]", "label": "性别", "field_key": "$2", "type": "cell"},
        {
            "path": "sheet[0]/row[27]/cell[1]",
            "label": "主要家庭成员",
            "field_key": "$3",
            "type": "dynamic_table",
            "table_config": {
                "dynamic": True,
                "header_row": 27,
                "data_start_row": 28,
                "data_end_row": 31,
                "columns": [
                    {"cell_index": 1, "name": "称谓"},
                    {"cell_index": 2, "name": "姓名"},
                    {"cell_index": 3, "name": "年龄"},
                    {"cell_index": 4, "name": "工作单位"},
                    {"cell_index": 5, "name": "职务"},
                    {"cell_index": 6, "name": "联系电话"},
                ],
            },
        },
    ]
    fields = [
        {"id": "$1", "value": "张明华"},
        {"id": "$2", "value": "男"},
        {"id": "$3", "rows": [{"称谓": f"成员{i}", "姓名": f"名{i}", "年龄": str(30 + i), "工作单位": f"单位{i}", "职务": "-", "联系电话": f"1380000000{i}"} for i in range(6)]},
    ]
    fill_workbook(REAL_SAMPLE, placeholders, fields, out)

    ws = openpyxl.load_workbook(out)["教职工信息登记表"]
    assert ws["C6"].value == "张明华"
    assert ws["B29"].value == "成员0" and ws["B34"].value == "成员5"  # 29-34 共6行(插2行)
    # 学习工作经历区整体下移2行: 原 B33 标题 -> B35
    assert ws["B35"].value.strip().startswith("七、")
    # 备注声明合并区 B41:G44 -> B43:G46
    merges = sorted(str(r) for r in ws.merged_cells.ranges)
    assert "B43:G46" in merges
    # 17 处下拉全部保留
    assert len(ws.data_validations.dataValidation) == 17
    # 其余 sheet 原样
    with zipfile.ZipFile(REAL_SAMPLE) as z1, zipfile.ZipFile(out) as z2:
        assert z1.read("xl/worksheets/sheet2.xml") == z2.read("xl/worksheets/sheet2.xml")


@pytest.mark.skipif(not os.path.exists(WPS_SAMPLE), reason="WPS 样例文件不存在")
def test_wps_sample_full_flow(tmp_path):
    """WPS 另存的文件：sharedStrings 单元格覆写、插行、WPS 特有部件保留"""
    wb = parse_xlsx(WPS_SAMPLE, os.path.basename(WPS_SAMPLE))
    main = wb.sheets[0]
    # WPS 把 ARGB 写成 FF alpha 前缀
    assert main.rows[5].cells[2].style.fill_color == "FFFFFDE7"
    assert main.rows[5].cells[4].validation_options == ["男", "女"]

    out = str(tmp_path / "filled.xlsx")
    placeholders = [
        {"path": "sheet[0]/row[5]/cell[2]", "label": "姓名", "field_key": "$1", "type": "cell"},
        {
            "path": "sheet[0]/row[27]/cell[1]",
            "label": "主要家庭成员",
            "field_key": "$2",
            "type": "dynamic_table",
            "table_config": {
                "dynamic": True,
                "header_row": 27,
                "data_start_row": 28,
                "data_end_row": 31,
                "columns": [{"cell_index": 1, "name": "称谓"}, {"cell_index": 2, "name": "姓名"}],
            },
        },
    ]
    fields = [
        {"id": "$1", "value": "张明华"},  # C6 原为 sharedStrings 空样式格
        {"id": "$2", "rows": [{"称谓": f"成员{i}", "姓名": f"名{i}"} for i in range(6)]},
    ]
    fill_workbook(WPS_SAMPLE, placeholders, fields, out)

    ws = openpyxl.load_workbook(out)["教职工信息登记表"]
    assert ws["C6"].value == "张明华"
    assert ws["B29"].value == "成员0" and ws["B34"].value == "成员5"
    assert len(ws.merged_cells.ranges) == 22
    assert len(ws.data_validations.dataValidation) == 17
    with zipfile.ZipFile(WPS_SAMPLE) as z1, zipfile.ZipFile(out) as z2:
        diff = [n for n in z1.namelist() if not n.endswith("/") and z1.read(n) != z2.read(n)]
        assert diff == ["xl/worksheets/sheet1.xml"]
        # WPS 特有部件原样保留
        assert z1.read("docProps/custom.xml") == z2.read("docProps/custom.xml")
        assert z1.read("xl/sharedStrings.xml") == z2.read("xl/sharedStrings.xml")
