"""
XLSX 结构解析（openpyxl 只读，不回写——回写走 filler 的 XML 直改，避免 openpyxl
save 重建文件导致形状/文本框等未建模元素丢失）
"""

import datetime
import logging

import openpyxl
from openpyxl.utils import get_column_letter

from .models import CellStyle, ParsedWorkbook, Sheet, SheetCell, SheetRow

logger = logging.getLogger(__name__)

# 表单类模板的合理上限，防止误传大数据表把接口打爆
MAX_ROWS = 2000
MAX_COLS = 100


def _display_value(value) -> str:
    """单元格显示值：日期统一 YYYY-MM-DD，整数值去掉小数点"""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _fill_color(cell) -> str | None:
    """提取纯 RGB 填充色；主题色（theme+tint）暂不换算，返回 None"""
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    color = fill.start_color
    if color is None or color.type != "rgb":
        return None
    rgb = color.rgb
    return rgb if isinstance(rgb, str) else None


def _font_color(cell) -> str | None:
    font = cell.font
    if font is None or font.color is None or font.color.type != "rgb":
        return None
    rgb = font.color.rgb
    return rgb if isinstance(rgb, str) else None


def _cell_style(cell) -> CellStyle:
    font = cell.font
    alignment = cell.alignment
    return CellStyle(
        fill_color=_fill_color(cell),
        bold=bool(font.bold) if font else False,
        font_size=float(font.size) if font and font.size else None,
        font_name=font.name if font else None,
        font_color=_font_color(cell),
        align_h=alignment.horizontal if alignment else None,
        align_v=alignment.vertical if alignment else None,
        wrap_text=bool(alignment.wrap_text) if alignment else False,
        number_format=cell.number_format if cell.number_format != "General" else None,
    )


def _merge_map(ws) -> dict[tuple[int, int], dict]:
    """
    构建合并信息映射：(row_idx, col_idx) 0-based -> merge 信息
    起始格记录 row_span/col_span，成员格记录 origin 坐标
    """
    result: dict[tuple[int, int], dict] = {}
    for rng in ws.merged_cells.ranges:
        origin = (rng.min_row - 1, rng.min_col - 1)
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                if (r, c) == origin:
                    result[(r, c)] = {
                        "origin": True,
                        "row_span": rng.max_row - rng.min_row + 1,
                        "col_span": rng.max_col - rng.min_col + 1,
                    }
                else:
                    result[(r, c)] = {"origin": False, "origin_pos": origin}
    return result


def _validation_map(ws) -> dict[tuple[int, int], list[str]]:
    """
    构建下拉校验映射：(row_idx, col_idx) 0-based -> 选项列表
    仅处理 type=list 且 formula1 为字面量（"a,b,c"）的校验；引用区域的暂不解析
    """
    result: dict[tuple[int, int], list[str]] = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue
        formula = dv.formula1.strip()
        if not (formula.startswith('"') and formula.endswith('"')):
            continue  # 引用单元格区域作为选项来源，首版不解析
        options = [opt.strip() for opt in formula[1:-1].split(",") if opt.strip()]
        if not options:
            continue
        for rng in dv.sqref.ranges:
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    result[(r, c)] = options
    return result


def parse_xlsx(file_path: str, filename: str) -> ParsedWorkbook:
    """
    解析 XLSX 文件为结构化模型（全部工作表）

    加载两次：data_only=True 取显示值（公式取缓存计算值），
    data_only=False 探测公式单元格。
    """
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    wb_formula = openpyxl.load_workbook(file_path, data_only=False)

    sheets: list[Sheet] = []
    for sheet_idx, ws in enumerate(wb_data.worksheets):
        ws_formula = wb_formula.worksheets[sheet_idx]
        max_row = min(ws.max_row or 1, MAX_ROWS)
        max_col = min(ws.max_column or 1, MAX_COLS)
        if (ws.max_row or 1) > MAX_ROWS or (ws.max_column or 1) > MAX_COLS:
            raise ValueError(f"工作表 [{ws.title}] 尺寸 {ws.max_row}x{ws.max_column} 超出模板上限 {MAX_ROWS}x{MAX_COLS}")

        merges = _merge_map(ws)
        validations = _validation_map(ws)
        sheet_path = f"sheet[{sheet_idx}]"

        rows: list[SheetRow] = []
        for r in range(max_row):
            row_path = f"{sheet_path}/row[{r}]"
            cells: list[SheetCell] = []
            for c in range(max_col):
                cell = ws.cell(row=r + 1, column=c + 1)
                formula_cell = ws_formula.cell(row=r + 1, column=c + 1)
                has_formula = formula_cell.data_type == "f" or (isinstance(formula_cell.value, str) and formula_cell.value.startswith("="))

                merge_info = merges.get((r, c))
                row_span, col_span, is_origin, origin_path = 1, 1, True, None
                if merge_info:
                    if merge_info["origin"]:
                        row_span = merge_info["row_span"]
                        col_span = merge_info["col_span"]
                    else:
                        is_origin = False
                        orow, ocol = merge_info["origin_pos"]
                        origin_path = f"{sheet_path}/row[{orow}]/cell[{ocol}]"

                cells.append(
                    SheetCell(
                        path=f"{row_path}/cell[{c}]",
                        value=_display_value(cell.value),
                        row_span=row_span,
                        col_span=col_span,
                        is_merged_origin=is_origin,
                        merge_origin_path=origin_path,
                        style=_cell_style(cell),
                        has_formula=has_formula,
                        validation_options=validations.get((r, c)),
                    )
                )
            height = ws.row_dimensions[r + 1].height if (r + 1) in ws.row_dimensions else None
            rows.append(SheetRow(path=row_path, height=height, cells=cells))

        col_widths: list[float | None] = []
        for c in range(max_col):
            letter = get_column_letter(c + 1)
            dim = ws.column_dimensions.get(letter)
            col_widths.append(dim.width if dim is not None and dim.width else None)

        sheets.append(
            Sheet(
                path=sheet_path,
                name=ws.title,
                index=sheet_idx,
                max_row=max_row,
                max_col=max_col,
                col_widths=col_widths,
                rows=rows,
            )
        )

    logger.info(f"[parse_xlsx] {filename}: {len(sheets)} 个工作表, 尺寸 {[f'{s.name}:{s.max_row}x{s.max_col}' for s in sheets]}")
    return ParsedWorkbook(filename=filename, sheets=sheets)
