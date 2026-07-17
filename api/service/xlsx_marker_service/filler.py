"""
XLSX 数据填充（XML 直改）

不使用 openpyxl 回写：openpyxl 的 save 会从对象模型重建整个文件，形状/文本框/
控件等未建模元素会静默丢失。此处将 xlsx 当作 zip 包，只修改目标工作表的
sheetN.xml（及发生插行时该表关联的 drawingN.xml 锚点），其余部件字节级原样拷贝。

坐标约定与 parser 一致：path = sheet[n]/row[m]/cell[k]，0-based。

插行（dynamic_table）采用自底向上处理：同一工作表内按锚点行号降序处理待填项，
插行只影响其下方内容，而下方内容此时已随 XML 节点整体平移完毕，坐标永不失效。
"""

import copy
import logging
import os
import re
import zipfile

import openpyxl
from lxml import etree
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

CELL_PATH_RE = re.compile(r"^sheet\[(\d+)\]/row\[(\d+)\]/cell\[(\d+)\]")
SHEET_PATH_RE = re.compile(r"^sheet\[(\d+)\]")
A1_RE = re.compile(r"([A-Z]+)(\d+)")


def _q(tag: str, ns: str = NS_MAIN) -> str:
    return f"{{{ns}}}{tag}"


def _a1(row_idx: int, col_idx: int) -> str:
    """0-based (row, col) -> A1 引用"""
    return f"{get_column_letter(col_idx + 1)}{row_idx + 1}"


def _col_index(letters: str) -> int:
    """列字母 -> 0-based 列索引"""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _shift_refs(ref_str: str, after_row: int, count: int) -> str:
    """把字符串中所有 A1 引用里行号 > after_row+1 的部分平移 count 行（after_row 为 0-based）"""
    boundary = after_row + 1  # 转 1-based

    def repl(m):
        row = int(m.group(2))
        return f"{m.group(1)}{row + count}" if row > boundary else m.group(0)

    return A1_RE.sub(repl, ref_str)


class SheetXml:
    """单个工作表 XML 的定向修改器"""

    def __init__(self, xml_bytes: bytes):
        self.root = etree.fromstring(xml_bytes)
        self.sheet_data = self.root.find(_q("sheetData"))
        if self.sheet_data is None:
            raise ValueError("工作表 XML 缺少 sheetData 节点")
        self.rows_inserted_total = 0

    def to_bytes(self) -> bytes:
        return etree.tostring(self.root, xml_declaration=True, encoding="UTF-8", standalone=True)

    # ---------- 写值 ----------

    def _find_row(self, row_idx: int):
        target = row_idx + 1
        for row in self.sheet_data.findall(_q("row")):
            r = int(row.get("r"))
            if r == target:
                return row
            if r > target:
                new_row = etree.Element(_q("row"), r=str(target))
                row.addprevious(new_row)
                return new_row
        new_row = etree.SubElement(self.sheet_data, _q("row"), r=str(target))
        return new_row

    def _find_cell(self, row_el, row_idx: int, col_idx: int):
        ref = _a1(row_idx, col_idx)
        for c in row_el.findall(_q("c")):
            m = A1_RE.match(c.get("r", ""))
            if m is None:
                continue
            if c.get("r") == ref:
                return c
            if _col_index(m.group(1)) > col_idx:
                new_c = etree.Element(_q("c"), r=ref)
                c.addprevious(new_c)
                return new_c
        new_c = etree.SubElement(row_el, _q("c"), r=ref)
        return new_c

    def set_cell(self, row_idx: int, col_idx: int, value: str):
        """写入单元格值（inlineStr），保留原样式属性 s=，清除公式/旧值"""
        row_el = self._find_row(row_idx)
        cell = self._find_cell(row_el, row_idx, col_idx)
        for child in list(cell):
            cell.remove(child)
        if value == "":
            cell.attrib.pop("t", None)
            return
        cell.set("t", "inlineStr")
        is_el = etree.SubElement(cell, _q("is"))
        t_el = etree.SubElement(is_el, _q("t"))
        t_el.text = value
        if value != value.strip() or "\n" in value:
            t_el.set(XML_SPACE, "preserve")

    # ---------- 插行 ----------

    def has_formula_below(self, after_row: int) -> bool:
        boundary = after_row + 1
        for row in self.sheet_data.findall(_q("row")):
            if int(row.get("r")) <= boundary:
                continue
            for c in row.findall(_q("c")):
                if c.find(_q("f")) is not None:
                    return True
        return False

    def insert_rows(self, after_row: int, count: int):
        """
        在 after_row（0-based）之后插入 count 行：
        克隆 after_row 行（含样式，清空值），重编号下方行/格，
        平移合并区/校验区/条件格式/超链接/autoFilter/dimension，
        克隆模板行内的横向合并到新行。
        """
        if count <= 0:
            return
        boundary = after_row + 1  # 1-based 模板行号

        template_row = None
        for row in self.sheet_data.findall(_q("row")):
            r = int(row.get("r"))
            if r == boundary:
                template_row = row
            elif r > boundary:
                row.set("r", str(r + count))
                for c in row.findall(_q("c")):
                    c.set("r", _shift_refs(c.get("r"), after_row, count))
        if template_row is None:
            raise ValueError(f"插行失败：模板行 {boundary} 在工作表 XML 中不存在（该行无任何样式/内容）")

        prev = template_row
        for i in range(1, count + 1):
            new_row = copy.deepcopy(template_row)
            new_row.set("r", str(boundary + i))
            for c in new_row.findall(_q("c")):
                m = A1_RE.match(c.get("r"))
                c.set("r", f"{m.group(1)}{boundary + i}")
                for child in list(c):
                    c.remove(child)
                c.attrib.pop("t", None)
            prev.addnext(new_row)
            prev = new_row

        self._shift_ranges(after_row, count)
        self._clone_row_merges(after_row, count)
        self.rows_inserted_total += count

    def _shift_ranges(self, after_row: int, count: int):
        merge_cells = self.root.find(_q("mergeCells"))
        if merge_cells is not None:
            for m_el in merge_cells.findall(_q("mergeCell")):
                m_el.set("ref", _shift_refs(m_el.get("ref"), after_row, count))
        dvs = self.root.find(_q("dataValidations"))
        if dvs is not None:
            for dv in dvs.findall(_q("dataValidation")):
                dv.set("sqref", _shift_refs(dv.get("sqref"), after_row, count))
        for cf in self.root.findall(_q("conditionalFormatting")):
            if cf.get("sqref"):
                cf.set("sqref", _shift_refs(cf.get("sqref"), after_row, count))
        hyperlinks = self.root.find(_q("hyperlinks"))
        if hyperlinks is not None:
            for link in hyperlinks.findall(_q("hyperlink")):
                link.set("ref", _shift_refs(link.get("ref"), after_row, count))
        auto_filter = self.root.find(_q("autoFilter"))
        if auto_filter is not None and auto_filter.get("ref"):
            auto_filter.set("ref", _shift_refs(auto_filter.get("ref"), after_row, count))
        dim = self.root.find(_q("dimension"))
        if dim is not None and dim.get("ref"):
            dim.set("ref", _shift_refs(dim.get("ref"), after_row, count))

    def _clone_row_merges(self, after_row: int, count: int):
        """模板行内的横向合并（如 F35:G35）在每个新行上克隆"""
        merge_cells = self.root.find(_q("mergeCells"))
        if merge_cells is None:
            return
        boundary = after_row + 1
        cloned = []
        for m_el in merge_cells.findall(_q("mergeCell")):
            ref = m_el.get("ref")
            m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", ref)
            if m is None:
                continue
            if int(m.group(2)) == boundary and int(m.group(4)) == boundary:
                for i in range(1, count + 1):
                    cloned.append(f"{m.group(1)}{boundary + i}:{m.group(3)}{boundary + i}")
        for ref in cloned:
            etree.SubElement(merge_cells, _q("mergeCell"), ref=ref)
        if cloned:
            merge_cells.set("count", str(len(merge_cells.findall(_q("mergeCell")))))

    def has_table_parts(self) -> bool:
        return self.root.find(_q("tableParts")) is not None

    def drawing_rel_id(self) -> str | None:
        drawing = self.root.find(_q("drawing"))
        return drawing.get(_q("id", NS_REL)) if drawing is not None else None


def _shift_drawing_anchors(xml_bytes: bytes, after_row: int, count: int) -> bytes:
    """平移绘图锚点：anchor 中的行号为 0-based，>= after_row+1 的下移"""
    root = etree.fromstring(xml_bytes)
    for anchor_tag in ("twoCellAnchor", "oneCellAnchor"):
        for anchor in root.findall(_q(anchor_tag, NS_XDR)):
            for pos_tag in ("from", "to"):
                pos = anchor.find(_q(pos_tag, NS_XDR))
                if pos is None:
                    continue
                row_el = pos.find(_q("row", NS_XDR))
                if row_el is not None and int(row_el.text) >= after_row + 1:
                    row_el.text = str(int(row_el.text) + count)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _workbook_sheet_parts(zf: zipfile.ZipFile) -> list[str]:
    """按 workbook 内顺序返回各工作表的部件路径，如 ['xl/worksheets/sheet1.xml', ...]"""
    wb_root = etree.fromstring(zf.read("xl/workbook.xml"))
    rels_root = etree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {rel.get("Id"): rel.get("Target") for rel in rels_root.findall(_q("Relationship", NS_PKG_REL))}

    parts = []
    sheets_el = wb_root.find(_q("sheets"))
    for sheet_el in sheets_el.findall(_q("sheet")):
        rid = sheet_el.get(_q("id", NS_REL))
        target = rel_map[rid]
        if target.startswith("/"):
            parts.append(target.lstrip("/"))
        else:
            parts.append(f"xl/{target}")
    return parts


def _sheet_rels_part(sheet_part: str) -> str:
    # xl/worksheets/sheet1.xml -> xl/worksheets/_rels/sheet1.xml.rels
    prefix, name = sheet_part.rsplit("/", 1)
    return f"{prefix}/_rels/{name}.rels"


def _drawing_part(zf: zipfile.ZipFile, sheet_part: str, rel_id: str) -> str | None:
    rels_name = _sheet_rels_part(sheet_part)
    if rels_name not in zf.namelist():
        return None
    rels_root = etree.fromstring(zf.read(rels_name))
    for rel in rels_root.findall(_q("Relationship", NS_PKG_REL)):
        if rel.get("Id") == rel_id:
            target = rel.get("Target")  # 一般为 ../drawings/drawing1.xml
            if target.startswith("/"):
                return target.lstrip("/")
            base = sheet_part.rsplit("/", 1)[0].split("/")
            for seg in target.split("/"):
                if seg == "..":
                    base.pop()
                else:
                    base.append(seg)
            return "/".join(base)
    return None


def _plan_writes(placeholders: list[dict], fields: list[dict]) -> dict[int, list[dict]]:
    """
    将 placeholders + fields 编排为按工作表分组的工作项：
    {sheet_idx: [{anchor_row, writes: [(row, col, value)], insert_after, insert_count}, ...]}
    """
    field_map = {f["id"]: f for f in fields}
    by_sheet: dict[int, list[dict]] = {}

    for ph in placeholders:
        field = field_map.get(ph.get("field_key"))
        if field is None:
            continue
        ph_type = ph.get("type", "cell")
        path = ph.get("path", "")

        if ph_type in ("cell", "summary"):
            m = CELL_PATH_RE.match(path)
            if m is None:
                raise ValueError(f"非法的单元格 path: {path}")
            sheet_idx, row_idx, col_idx = int(m.group(1)), int(m.group(2)), int(m.group(3))
            value = str(field.get("value", "") or "")
            item = {"anchor_row": row_idx, "writes": [(row_idx, col_idx, value)], "insert_after": None, "insert_count": 0}
            by_sheet.setdefault(sheet_idx, []).append(item)
            continue

        if ph_type in ("table", "dynamic_table"):
            m = SHEET_PATH_RE.match(path)
            cfg = ph.get("table_config")
            if m is None or not cfg:
                raise ValueError(f"表格待填项缺少合法 path 或 table_config: {path}")
            sheet_idx = int(m.group(1))
            data_start, data_end = int(cfg["data_start_row"]), int(cfg["data_end_row"])
            capacity = data_end - data_start + 1
            columns = {col["name"]: int(col["cell_index"]) for col in cfg.get("columns", [])}
            rows_data = field.get("rows") or []

            dynamic = ph_type == "dynamic_table" or cfg.get("dynamic")
            if not dynamic and len(rows_data) > capacity:
                logger.warning(f"[fill] 固定表格 {ph.get('field_key')} 数据 {len(rows_data)} 行超出预留 {capacity} 行，截断")
                rows_data = rows_data[:capacity]

            insert_count = max(0, len(rows_data) - capacity) if dynamic else 0
            writes = []
            for i, row_data in enumerate(rows_data):
                target_row = data_start + i
                for col_name, value in row_data.items():
                    if col_name not in columns:
                        logger.warning(f"[fill] 表格 {ph.get('field_key')} 出现未配置列名 [{col_name}]，忽略")
                        continue
                    writes.append((target_row, columns[col_name], str(value or "")))
            item = {
                "anchor_row": data_start,
                "writes": writes,
                "insert_after": data_end if insert_count > 0 else None,
                "insert_count": insert_count,
            }
            by_sheet.setdefault(sheet_idx, []).append(item)
            continue

        raise ValueError(f"不支持的待填项类型: {ph_type}")

    return by_sheet


def fill_workbook(source_path: str, placeholders: list[dict], fields: list[dict], output_path: str):
    """
    填充入口：XML 直改，写后自校验。

    - placeholders/fields 结构与 docx_marker 同构（见 models.Placeholder / models.FieldValue）
    - 固定表格超出预留行截断；动态表格自动插行
    - 任何校验失败直接抛错，绝不产出可疑文件
    """
    by_sheet = _plan_writes(placeholders, fields)
    expected: list[tuple[int, int, int, str]] = []  # (sheet, row, col, value) 用于写后校验，坐标为插行后的最终坐标

    with zipfile.ZipFile(source_path) as zf:
        sheet_parts = _workbook_sheet_parts(zf)
        all_names = zf.namelist()

        patched: dict[str, bytes] = {}
        for sheet_idx, items in by_sheet.items():
            if sheet_idx >= len(sheet_parts):
                raise ValueError(f"path 引用了不存在的工作表索引 sheet[{sheet_idx}]")
            part = sheet_parts[sheet_idx]
            sheet = SheetXml(zf.read(part))

            if sheet.has_table_parts() and any(it["insert_count"] > 0 for it in items):
                raise ValueError("该工作表包含结构化表格（tableParts），暂不支持动态插行")

            # 自底向上：按锚点行号降序处理，插行只影响已处理完的下方内容
            # （下方内容随 XML 节点整体平移，坐标不会失效；但校验清单里已记录的
            #   下方坐标需要同步平移，否则写后自校验会拿旧坐标误报）
            items.sort(key=lambda it: it["anchor_row"], reverse=True)
            sheet_expected: list[list] = []
            for item in items:
                if item["insert_count"] > 0:
                    if sheet.has_formula_below(item["insert_after"]):
                        raise ValueError(f"动态表格下方存在公式单元格，暂不支持插行（插入点行 {item['insert_after'] + 1}）")
                    sheet.insert_rows(item["insert_after"], item["insert_count"])
                    for entry in sheet_expected:
                        if entry[0] > item["insert_after"]:
                            entry[0] += item["insert_count"]
                for row_idx, col_idx, value in item["writes"]:
                    sheet.set_cell(row_idx, col_idx, value)
                    sheet_expected.append([row_idx, col_idx, value])
            expected.extend((sheet_idx, r, c, v) for r, c, v in sheet_expected)

            patched[part] = sheet.to_bytes()

            # 发生插行且该表挂有绘图 -> 平移锚点
            inserted_points = [(it["insert_after"], it["insert_count"]) for it in items if it["insert_count"] > 0]
            if inserted_points:
                rel_id = sheet.drawing_rel_id()
                if rel_id:
                    drawing_part = _drawing_part(zf, part, rel_id)
                    if drawing_part and drawing_part in all_names:
                        drawing_bytes = patched.get(drawing_part, zf.read(drawing_part))
                        for after_row, count in inserted_points:  # 已按自底向上排序，逐次平移
                            drawing_bytes = _shift_drawing_anchors(drawing_bytes, after_row, count)
                        patched[drawing_part] = drawing_bytes

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zf.infolist():
                zout.writestr(item, patched.get(item.filename, zf.read(item.filename)))

    _verify_output(output_path, expected)
    logger.info(f"[fill_workbook] 完成: {len(expected)} 个单元格写入, 输出 {output_path}")


def _verify_output(output_path: str, expected: list[tuple[int, int, int, str]]):
    """写后自校验：文件可打开 + 逐格读回比对。失败即抛错，不产出可疑文件。"""
    try:
        wb = openpyxl.load_workbook(output_path, data_only=False)
    except Exception as e:
        os.remove(output_path)
        raise RuntimeError(f"填充自校验失败：输出文件无法打开（{e!s}），已丢弃") from e

    mismatches = []
    for sheet_idx, row_idx, col_idx, value in expected:
        actual = wb.worksheets[sheet_idx].cell(row=row_idx + 1, column=col_idx + 1).value
        actual_str = "" if actual is None else str(actual)
        if actual_str != value:
            mismatches.append(f"sheet[{sheet_idx}] {_a1(row_idx, col_idx)}: 期望[{value}] 实际[{actual_str}]")
    if mismatches:
        os.remove(output_path)
        raise RuntimeError("填充自校验失败，已丢弃输出文件: " + "; ".join(mismatches[:5]))
